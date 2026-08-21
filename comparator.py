"""
comparator.py
-------------
Excel loading + ledger cleaning utilities.

This app has one job: take a messy "Rincian Buku Besar" style export and
keep only the real transaction rows, identified by the font (name + size)
used on the cell in the 'Tanggal' column. Page headers, footers, and
repeated header rows all get dropped.
"""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
#  Basic utilities
# ------------------------------------------------------------
def load_workbook_safe(path):
    # data_only=False so we keep formulas as-is; styles (fonts) always load
    # regardless of this flag.
    return openpyxl.load_workbook(path, data_only=False)


def sheet_names(wb):
    return wb.sheetnames


# ------------------------------------------------------------
#  Header detection helpers
# ------------------------------------------------------------
HEADER_TERMS = ["Tanggal", "Tipe Transaksi", "Keterangan", "Debit", "Kredit", "Saldo Akhir"]


def is_header_row(ws, row_idx):
    """Return True if the row contains all header terms (case-insensitive)."""
    row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
    row_text = " ".join(row_values).lower()
    return all(term.lower() in row_text for term in HEADER_TERMS)


def find_header_row(ws, header_terms=None):
    if header_terms is None:
        header_terms = HEADER_TERMS
    terms_lower = [t.lower() for t in header_terms]
    for row_idx in range(1, min(31, ws.max_row + 1)):
        row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
        row_text = " ".join(row_values).lower()
        if all(term in row_text for term in terms_lower):
            return row_idx
    return None


# ------------------------------------------------------------
#  Ledger cleaning – font-based method
# ------------------------------------------------------------
def clean_ledger_by_formatting(source_wb, sheet_name, output_path, font_name="Arial", font_size=9):
    """
    Create a new Excel workbook containing only rows where the cell in the
    'Tanggal' column uses the specified font and size.
    Duplicate header rows are automatically skipped.

    Returns (new_workbook, kept_row_count_excluding_header).
    """
    ws = source_wb[sheet_name]

    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError("Baris header tidak ditemukan pada sheet ini.")

    col_map = {
        'tanggal': 4, 'tipe transaksi': 6, 'keterangan': 9,
        'debit': 11, 'kredit': 13, 'saldo akhir': 15
    }
    header_cells = {}
    for col_idx in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col_idx).value or "").strip()
        if val:
            header_cells[val.lower()] = col_idx
    for key in col_map.keys():
        if key in header_cells:
            col_map[key] = header_cells[key]

    date_col = col_map['tanggal']
    keep_rows = [header_row]  # always keep the first header

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Skip duplicate header rows early
        if is_header_row(ws, row_idx):
            continue

        cell = ws.cell(row=row_idx, column=date_col)
        font = cell.font
        if font is None:
            continue
        name_match = font.name and font.name.lower() == font_name.lower()
        size_match = font.size and abs(font.size - font_size) < 0.5
        if name_match and size_match:
            keep_rows.append(row_idx)

    if len(keep_rows) <= 1:
        raise ValueError(
            f"Tidak ada baris yang cocok dengan font '{font_name}' ukuran {font_size}. "
            "Coba sesuaikan pengaturan font."
        )

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if ws.column_dimensions[col_letter].width:
            new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

    def copy_row(src_ws, src_row_num, dest_ws, dest_row_num):
        if src_ws.row_dimensions[src_row_num].height:
            dest_ws.row_dimensions[dest_row_num].height = src_ws.row_dimensions[src_row_num].height
        for col_idx in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=src_row_num, column=col_idx)
            dest_cell = dest_ws.cell(row=dest_row_num, column=col_idx)
            dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = src_cell.font.copy()
                dest_cell.border = src_cell.border.copy() if src_cell.border else None
                dest_cell.fill = src_cell.fill.copy() if src_cell.fill else None
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = src_cell.alignment.copy() if src_cell.alignment else None

    dest_row = 1
    for src_row in keep_rows:
        copy_row(ws, src_row, new_ws, dest_row)
        dest_row += 1

    new_wb.save(output_path)
    return new_wb, len(keep_rows) - 1
