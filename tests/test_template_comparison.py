import os
import tempfile
import unittest
from unittest.mock import patch
from openpyxl import Workbook

from app import (
    build_header_grid,
    detect_skeleton_from_workbook,
    build_reference_grid_from_paths,
    find_matching_skeleton_for_paths,
    diff_workbooks_by_header,
    build_directional_comparisons,
)


class TemplateComparisonTests(unittest.TestCase):
    def test_detect_skeleton_from_workbook_content(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, "template.xlsx")
            template_wb = Workbook()
            template_ws = template_wb.active
            template_ws.title = "Sheet1"
            template_ws["B1"] = "Col A"
            template_ws["C1"] = "Col B"
            template_ws["A2"] = "Row 1"
            template_wb.save(template_path)

            detected = detect_skeleton_from_workbook(template_path)
            self.assertTrue(detected)

    def test_build_reference_grid_from_paths_uses_shared_headers(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base_path = os.path.join(tmpdir, "base.xlsx")
            compare_path = os.path.join(tmpdir, "compare.xlsx")
            template_path = os.path.join(tmpdir, "template.xlsx")

            for path, headers, rows in [
                (base_path, ["Col A", "Col B", "Col C"], ["Row 1", "Row 2"]),
                (compare_path, ["Col A", "Col B", "Col D"], ["Row 1", "Row 3"]),
                (template_path, ["Col A", "Col B", "Col E"], ["Row 1", "Row 2"]),
            ]:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                ws["A1"] = "Row"
                for idx, header in enumerate(headers, start=2):
                    ws.cell(row=1, column=idx, value=header)
                for row_idx, row_name in enumerate(rows, start=2):
                    ws.cell(row=row_idx, column=1, value=row_name)
                wb.save(path)

            reference_grid = build_reference_grid_from_paths([base_path, compare_path, template_path], "Sheet1")
            self.assertEqual(set(reference_grid["col_headers"]), {"Col A", "Col B"})
            self.assertEqual(set(reference_grid["row_headers"]), {"Row 1"})

    def test_build_header_grid_uses_reference_headers(self):
        template_wb = Workbook()
        template_ws = template_wb.active
        template_ws.title = "Sheet1"
        template_ws["B1"] = "Col A"
        template_ws["C1"] = "Col B"
        template_ws["A2"] = "Row 1"
        template_ws["B2"] = 10
        template_ws["C2"] = 20

        reference_grid = build_header_grid(template_ws)

        actual_wb = Workbook()
        actual_ws = actual_wb.active
        actual_ws.title = "Sheet1"
        actual_ws["B1"] = "Col B"
        actual_ws["C1"] = "Col A"
        actual_ws["A2"] = "Row 1"
        actual_ws["B2"] = 20
        actual_ws["C2"] = 10

        actual_grid = build_header_grid(actual_ws, reference_grid=reference_grid)

        self.assertEqual(actual_grid["col_headers"], ["Col A", "Col B"])
        self.assertEqual(actual_grid["row_headers"], ["Row 1"])
        self.assertEqual(actual_grid["data"][("Row 1", "Col A")], 10)
        self.assertEqual(actual_grid["data"][("Row 1", "Col B")], 20)

    def test_find_matching_skeleton_for_paths_uses_existing_skeleton(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            skeleton_dir = os.path.join(tmpdir, "skeletons")
            os.makedirs(skeleton_dir, exist_ok=True)

            skeleton_path = os.path.join(skeleton_dir, "template.xlsx")
            skeleton_wb = Workbook()
            skeleton_ws = skeleton_wb.active
            skeleton_ws.title = "Sheet1"
            skeleton_ws["B1"] = "Col A"
            skeleton_ws["C1"] = "Col B"
            skeleton_ws["A2"] = "Row 1"
            skeleton_wb.save(skeleton_path)

            compare_path = os.path.join(tmpdir, "compare.xlsx")
            compare_wb = Workbook()
            compare_ws = compare_wb.active
            compare_ws.title = "Sheet1"
            compare_ws["B1"] = "Col A"
            compare_ws["C1"] = "Col B"
            compare_ws["A2"] = "Row 1"
            compare_ws["B2"] = 42
            compare_wb.save(compare_path)

            with patch("app.SKELETON_DIR", skeleton_dir):
                matched_path = find_matching_skeleton_for_paths([compare_path])

            self.assertEqual(matched_path, skeleton_path)

    def test_diff_workbooks_by_header_preserves_reference_headers_for_all_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base_path = os.path.join(tmpdir, "base.xlsx")
            compare_path = os.path.join(tmpdir, "compare.xlsx")
            template_path = os.path.join(tmpdir, "template.xlsx")

            template_wb = Workbook()
            template_ws = template_wb.active
            template_ws.title = "Sheet1"
            template_ws["B1"] = "Col A"
            template_ws["C1"] = "Col B"
            template_ws["A2"] = "Row 1"
            template_ws2 = template_wb.create_sheet("Sheet2")
            template_ws2["B1"] = "Col X"
            template_ws2["C1"] = "Col Y"
            template_ws2["A2"] = "Row 1"
            template_wb.save(template_path)

            base_wb = Workbook()
            base_ws = base_wb.active
            base_ws.title = "Sheet1"
            base_ws["B1"] = "Col A"
            base_ws["C1"] = "Col B"
            base_ws["A2"] = "Row 1"
            base_ws2 = base_wb.create_sheet("Sheet2")
            base_ws2["B1"] = "Col X"
            base_ws2["C1"] = "Col Y"
            base_ws2["A2"] = "Row 1"
            base_wb.save(base_path)

            compare_wb = Workbook()
            compare_ws = compare_wb.active
            compare_ws.title = "Sheet1"
            compare_ws["B1"] = "Col A"
            compare_ws["C1"] = "Col B"
            compare_ws["A2"] = "Row 1"
            compare_ws2 = compare_wb.create_sheet("Sheet2")
            compare_ws2["B1"] = "Col X"
            compare_ws2["C1"] = "Col Y"
            compare_ws2["A2"] = "Row 1"
            compare_wb.save(compare_path)

            result = diff_workbooks_by_header(base_path, compare_path, reference_path=template_path)

            self.assertEqual(result["sheets"]["Sheet1"]["reference_col_headers"], ["Col A", "Col B"])
            self.assertEqual(result["sheets"]["Sheet2"]["reference_col_headers"], ["Col X", "Col Y"])

    def test_build_header_grid_uses_detected_header_row(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Title"
        ws["B1"] = "Ignored"
        ws["A2"] = "Row"
        ws["B2"] = "ABC"
        ws["C2"] = "DEF"
        ws["A3"] = "Item"
        ws["B3"] = 10
        ws["C3"] = 20

        grid = build_header_grid(ws)

        self.assertEqual(grid["col_headers"], ["ABC", "DEF"])
        self.assertEqual(grid["row_headers"], ["Item"])
        self.assertEqual(grid["data"][("Item", "ABC")], 10)
        self.assertEqual(grid["data"][("Item", "DEF")], 20)

    def test_build_directional_comparisons_returns_two_directions(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            file1_path = os.path.join(tmpdir, "file1.xlsx")
            file2_path = os.path.join(tmpdir, "file2.xlsx")

            wb1 = Workbook()
            ws1 = wb1.active
            ws1.title = "Sheet1"
            ws1["B1"] = "Col A"
            ws1["C1"] = "Col B"
            ws1["A2"] = "Row 1"
            wb1.save(file1_path)

            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "Sheet1"
            ws2["B1"] = "Col A"
            ws2["C1"] = "Col B"
            ws2["A2"] = "Row 1"
            wb2.save(file2_path)

            comparisons = build_directional_comparisons(file1_path, file2_path, "file1.xlsx", "file2.xlsx")

            self.assertEqual(len(comparisons), 2)
            self.assertEqual(comparisons[0]["direction"], "file1_to_file2")
            self.assertEqual(comparisons[1]["direction"], "file2_to_file1")

    def test_compare_sheet_grids_omits_empty_columns(self):
        from app import compare_sheet_grids

        grid1 = {
            "col_headers": ["Col A", "Col B"],
            "row_headers": ["Row 1"],
            "row_pos": {"Row 1": 2},
            "col_pos": {"Col A": 2, "Col B": 3},
            "data": {("Row 1", "Col A"): 1},
        }
        grid2 = {
            "col_headers": ["Col A", "Col B", "Col C"],
            "row_headers": ["Row 1"],
            "row_pos": {"Row 1": 2},
            "col_pos": {"Col A": 2, "Col B": 3, "Col C": 4},
            "data": {("Row 1", "Col A"): 1, ("Row 1", "Col B"): None, ("Row 1", "Col C"): None},
        }

        result = compare_sheet_grids(grid1, grid2)

        self.assertEqual(result["col_headers"], ["Col A"])


if __name__ == "__main__":
    unittest.main()
