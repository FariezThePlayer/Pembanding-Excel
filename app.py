import os
import io
import re
import time
import uuid
import shutil
import tempfile
import datetime
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import load_workbook
from werkzeug.exceptions import HTTPException

app = Flask(__name__)
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

# Semua file yang diunggah pengguna HANYA disimpan sementara di folder temp
# sistem operasi (bukan di dalam folder proyek) dan dihapus otomatis:
#   - setiap kali ada request baru, job yang sudah lebih tua dari
#     JOB_TTL_SECONDS akan disapu dan dihapus (lihat cleanup_expired_jobs), dan
#   - setiap kali job dipakai (compare/merge), "jam" job tersebut di-reset
#     supaya sesi yang masih aktif dipakai tidak keburu terhapus.
# Tidak ada file unggahan yang disimpan permanen.
JOBS_DIR = os.path.join(tempfile.gettempdir(), "excel_toolkit_jobs")
os.makedirs(JOBS_DIR, exist_ok=True)
JOB_TTL_SECONDS = 30 * 60  # 30 menit tidak dipakai -> dihapus otomatis

# Folder skeleton/template opsional. Ini BUKAN tempat menyimpan file yang
# diunggah pengguna — folder ini hanya dipakai untuk MENCOCOKKAN terhadap
# template yang sudah ada di sana (mis. ditaruh manual oleh admin). Aplikasi
# tidak lagi menulis/menyimpan file unggahan ke folder ini.
SKELETON_DIR = os.path.join(os.path.dirname(__file__), "skeletons")
os.makedirs(SKELETON_DIR, exist_ok=True)

# Penanda internal untuk sel yang isinya rumus/fungsi Excel (mis. =SUM(A1:A5)).
# Nilai hasil rumus TIDAK dibandingkan/digabung — sel seperti ini akan
# ditampilkan sebagai "(fungsi excel)" di UI.
FORMULA_MARKER = "@@EXCEL_FORMULA@@"
FORMULA_DISPLAY = "(Formula Excel)"


def cleanup_expired_jobs():
    """Hapus folder job (hasil unggahan sementara) yang sudah lebih tua dari JOB_TTL_SECONDS."""
    now = time.time()
    try:
        entries = os.listdir(JOBS_DIR)
    except OSError:
        return
    for entry in entries:
        full_path = os.path.join(JOBS_DIR, entry)
        try:
            age = now - os.path.getmtime(full_path)
        except OSError:
            continue
        if age > JOB_TTL_SECONDS:
            shutil.rmtree(full_path, ignore_errors=True)


def touch_job(job_id):
    """Perbarui waktu 'terakhir dipakai' sebuah job supaya tidak keburu dihapus otomatis."""
    job_dir = os.path.join(JOBS_DIR, job_id)
    try:
        os.utime(job_dir, None)
    except OSError:
        pass


def col_letter(idx):
    """Nomor kolom (1-indexed) -> huruf kolom Excel."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def safe_value(v):
    """Ubah value openpyxl menjadi sesuatu yang aman untuk JSON."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    return v


def is_formula_cell(cell):
    if cell is None:
        return False
    if getattr(cell, "data_type", None) == "f":
        return True
    return isinstance(cell.value, str) and str(cell.value).startswith("=")


def normalize_cell_value(cell):
    if cell is None:
        return None
    if is_formula_cell(cell):
        return FORMULA_DISPLAY
    return safe_value(cell.value)


def normalize_header_label(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s*\[[^\]]*\]\s*", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    if not text:
        return ""
    tokens = text.split()
    if "tarif" in tokens or "tariff" in tokens:
        return "tarif"
    return " ".join(tokens)


def format_header_label(label, column_idx=None):
    if label is None:
        label = ""
    text = str(label).strip()
    if not text:
        return ""
    if column_idx is None:
        return text
    return f"{text} [{col_letter(column_idx)}]"


def column_has_values(ws, col_idx, header_row, max_row):
    for row_idx in range(header_row, max_row + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value not in (None, ""):
            return True
    return False


def detect_header_row(ws, max_scan=10):
    """
    Cari baris yang paling mungkin jadi baris header kolom. Untuk layout yang
    tidak standar, kita cari baris yang paling banyak mengandung label teks
    dan yang paling dekat ke bagian atas sheet.
    """
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0
    if max_row == 0 or max_col < 2:
        return 1

    scan_limit = min(max_row, max_scan)
    best_row = 1
    best_score = -1

    for r in range(1, scan_limit + 1):
        filled_cells = 0
        text_like_cells = 0
        for c in range(2, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is None or str(val).strip() == "":
                continue
            filled_cells += 1
            text = str(val).strip()
            digits_only = text.replace(" ", "").replace("-", "").replace(".", "")
            if not (digits_only.isdigit() and len(digits_only) >= 6):
                text_like_cells += 1

        if filled_cells >= 2 and text_like_cells > best_score:
            best_score = text_like_cells
            best_row = r

    return best_row


def detect_skeleton_from_workbook(path):
    try:
        wb = load_workbook(path, data_only=False)
    except Exception:
        return False

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_col <= 1:
            continue

        header_row = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        first_col = [ws.cell(row=r, column=1).value for r in range(1, max_row + 1)]

        header_values = [str(v).strip() for v in header_row[1:4] if v not in (None, "")]
        row_values = [str(v).strip() for v in first_col[1:4] if v not in (None, "")]
        has_header_like_row = bool(header_values)
        has_header_like_col = bool(row_values)

        non_empty_cells = 0
        for r in range(2, min(max_row, 6) + 1):
            for c in range(2, min(max_col, 6) + 1):
                value = ws.cell(row=r, column=c).value
                if value not in (None, ""):
                    non_empty_cells += 1

        if has_header_like_row and has_header_like_col and non_empty_cells <= 3:
            return True

    return False


def fingerprint_workbook(path):
    try:
        wb = load_workbook(path, data_only=False)
    except Exception:
        return None

    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            parts.append(tuple(row[:min(ws.max_column, 8)]))

    return str(parts).encode("utf-8")


def find_matching_skeleton_for_paths(paths):
    """
    Cocokkan file yang diunggah terhadap library skeleton/template yang ADA
    di SKELETON_DIR (mis. ditaruh manual oleh admin). Fungsi ini hanya
    MEMBACA folder tersebut — tidak pernah menulis file unggahan ke sana.
    """
    if not os.path.isdir(SKELETON_DIR):
        return None

    candidates = []
    for entry in os.listdir(SKELETON_DIR):
        full_path = os.path.join(SKELETON_DIR, entry)
        if not os.path.isfile(full_path):
            continue
        try:
            wb = load_workbook(full_path, data_only=False)
        except Exception:
            continue
        candidates.append((full_path, wb))

    if not candidates:
        return None

    best_match = None
    best_score = -1
    for candidate_path, candidate_wb in candidates:
        score = 0
        for path in paths:
            if not os.path.exists(path):
                continue
            try:
                wb = load_workbook(path, data_only=False)
            except Exception:
                continue

            for sheet_name in set(wb.sheetnames) & set(candidate_wb.sheetnames):
                ws1 = wb[sheet_name]
                ws2 = candidate_wb[sheet_name]

                header_row1 = detect_header_row(ws1)
                header_row2 = detect_header_row(ws2)

                headers1 = [normalize_header_label(ws1.cell(row=header_row1, column=c).value) for c in range(2, ws1.max_column + 1)]
                headers2 = [normalize_header_label(ws2.cell(row=header_row2, column=c).value) for c in range(2, ws2.max_column + 1)]

                overlap = len(set(headers1) & set(headers2))
                score += overlap

                row_headers1 = [normalize_header_label(ws1.cell(row=r, column=1).value) for r in range(header_row1 + 1, ws1.max_row + 1)]
                row_headers2 = [normalize_header_label(ws2.cell(row=r, column=1).value) for r in range(header_row2 + 1, ws2.max_row + 1)]
                score += len(set(row_headers1) & set(row_headers2))

        if score > best_score:
            best_score = score
            best_match = candidate_path

    if best_score <= 0:
        return None
    return best_match

def natural_sort_key(label):
    """Sort key yang memperlakukan angka di dalam teks secara numerik,
    supaya 'Baris 5' < 'Baris 6' < ... < 'Baris 10' (bukan urut alfabet)."""
    return [int(chunk) if chunk.isdigit() else chunk.lower() for chunk in re.split(r"(\d+)", str(label))]

def build_reference_grid_from_paths(paths, sheet_name):
    grids = []
    for path in paths:
        if not os.path.exists(path):
            continue
        try:
            wb = load_workbook(path, data_only=False)
        except Exception:
            continue
        if sheet_name not in wb.sheetnames:
            continue
        grids.append(build_header_grid(wb[sheet_name]))

    if not grids:
        return None

    shared_col_headers = None
    shared_row_headers = None
    col_display_map = {}
    row_display_map = {}

    for grid in grids:
        grid_col_headers = list(grid["col_headers"])
        grid_row_headers = list(grid["row_headers"])

        for label in grid_col_headers:
            norm = normalize_header_label(label)
            if norm not in col_display_map:
                col_display_map[norm] = label

        for label in grid_row_headers:
            norm = normalize_header_label(label)
            if norm not in row_display_map:
                row_display_map[norm] = label

        normalized_cols = {normalize_header_label(label) for label in grid_col_headers}
        normalized_rows = {normalize_header_label(label) for label in grid_row_headers}

        if shared_col_headers is None:
            shared_col_headers = normalized_cols
        else:
            shared_col_headers &= normalized_cols

        if shared_row_headers is None:
            shared_row_headers = normalized_rows
        else:
            shared_row_headers &= normalized_rows

    if shared_col_headers is None:
        shared_col_headers = set()
    if shared_row_headers is None:
        shared_row_headers = set()

    return {
        "col_headers": [col_display_map[norm] for norm in sorted(shared_col_headers, key=natural_sort_key)],
        "row_headers": [row_display_map[norm] for norm in sorted(shared_row_headers, key=natural_sort_key)],
    }


def build_header_grid(ws, reference_grid=None):
    """
    Membaca sebuah sheet dan menganggap:
      - baris header  = baris yang terdeteksi lewat detect_header_row()
      - kolom A       = header baris
    Mengembalikan dict berisi urutan header, posisi asli header di sheet
    (untuk keperluan penulisan ulang saat merge), dan data yang sudah
    diindeks berdasarkan (header_baris, header_kolom) alih-alih posisi sel
    mentah. Ini membuat perbandingan tetap benar walau baris/kolom tertukar
    urutan, ditambah, atau dihapus di salah satu file — dan tahan terhadap
    file yang header-nya tidak persis di baris 1.

    Saat reference_grid disediakan, header/row label dari template tersebut
    dipakai sebagai struktur acuan; file yang dibandingkan akan dibaca sesuai
    label-label tersebut.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    header_row = detect_header_row(ws)

    if reference_grid is not None:
        col_headers = list(reference_grid["col_headers"])
        row_headers = list(reference_grid["row_headers"])
        col_pos = {}
        row_pos = {}

        for c_label in col_headers:
            for c in range(2, max_col + 1):
                raw = ws.cell(row=header_row, column=c).value
                label = str(raw).strip() if raw is not None and str(raw).strip() != "" else f"Kolom {col_letter(c)}"
                if normalize_header_label(label) == normalize_header_label(c_label):
                    col_pos[c_label] = c
                    break

        for r_label in row_headers:
            for r in range(header_row + 1, max_row + 1):
                raw = ws.cell(row=r, column=1).value
                label = str(raw).strip() if raw is not None and str(raw).strip() != "" else f"Baris {r}"
                if normalize_header_label(label) == normalize_header_label(r_label):
                    row_pos[r_label] = r
                    break

        data = {}
        for r_label, r in row_pos.items():
            for c_label, c in col_pos.items():
                cell = ws.cell(row=r, column=c)
                data[(r_label, c_label)] = normalize_cell_value(cell)

        return {
            "col_headers": col_headers,
            "row_headers": row_headers,
            "row_pos": row_pos,
            "col_pos": col_pos,
            "data": data,
            "header_row": header_row,
        }

    col_headers = []
    col_pos = {}
    seen_col = {}
    display_col_headers = []
    for c in range(2, max_col + 1):
        if not column_has_values(ws, c, header_row, max_row):
            continue
        raw = ws.cell(row=header_row, column=c).value
        label = str(raw).strip() if raw is not None and str(raw).strip() != "" else f"Kolom {col_letter(c)}"
        if label in seen_col:
            seen_col[label] += 1
            label = f"{label} ({seen_col[label]})"
        else:
            seen_col[label] = 1
        col_headers.append(label)
        col_pos[label] = c
        display_col_headers.append(format_header_label(label, c))

    row_headers = []
    row_pos = {}
    seen_row = {}
    for r in range(header_row + 1, max_row + 1):
        raw = ws.cell(row=r, column=1).value
        label = str(raw).strip() if raw is not None and str(raw).strip() != "" else f"Baris {r}"
        if label in seen_row:
            seen_row[label] += 1
            label = f"{label} ({seen_row[label]})"
        else:
            seen_row[label] = 1
        row_headers.append(label)
        row_pos[label] = r

    if not row_headers:
        for r in range(header_row + 1, max_row + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(2, max_col + 1)):
                row_label = ws.cell(row=r, column=1).value
                if row_label not in (None, ""):
                    label = str(row_label).strip()
                    row_headers.append(label)
                    row_pos[label] = r

    if not col_headers:
        for c in range(2, max_col + 1):
            header_value = ws.cell(row=header_row, column=c).value
            if header_value not in (None, ""):
                label = str(header_value).strip()
                col_headers.append(label)
                col_pos[label] = c

    data = {}
    for r_label, r in row_pos.items():
        for c_label, c in col_pos.items():
            cell = ws.cell(row=r, column=c)
            data[(r_label, c_label)] = normalize_cell_value(cell)

    return {
        "col_headers": col_headers,
        "display_col_headers": display_col_headers,
        "row_headers": row_headers,
        "row_pos": row_pos,
        "col_pos": col_pos,
        "data": data,
        "header_row": header_row,
    }


def compare_sheet_grids(grid1, grid2, reference_grid=None):
    """Bandingkan dua grid (sudah diindeks berdasarkan header) dari sheet yang sama."""
    if reference_grid is not None:
        col_headers = list(reference_grid.get("col_headers", []))
        row_headers = list(reference_grid.get("row_headers", []))
    else:
        col_headers = list(grid1["col_headers"])
        col_set = set(col_headers)
        for c in grid2["col_headers"]:
            if c not in col_set:
                col_headers.append(c)
                col_set.add(c)

        row_headers = list(grid1["row_headers"])
        row_set = set(row_headers)
        for r in grid2["row_headers"]:
            if r not in row_set:
                row_headers.append(r)
                row_set.add(r)

    visible_col_headers = []
    display_col_headers = []
    for c_label in col_headers:
        has_any_value = False
        for r_label in row_headers:
            has1 = r_label in grid1["row_pos"] and c_label in grid1["col_pos"]
            has2 = r_label in grid2["row_pos"] and c_label in grid2["col_pos"]
            if not has1 and not has2:
                continue
            v1 = grid1["data"].get((r_label, c_label)) if has1 else None
            v2 = grid2["data"].get((r_label, c_label)) if has2 else None
            if v1 not in (None, "") or v2 not in (None, ""):
                has_any_value = True
                break
        if has_any_value:
            visible_col_headers.append(c_label)
            col_idx = (grid1.get("col_pos", {}).get(c_label) or grid2.get("col_pos", {}).get(c_label))
            display_col_headers.append(format_header_label(c_label, col_idx))

    rows_out = []
    total_diff = 0

    for r_label in row_headers:
        has_row1 = r_label in grid1["row_pos"]
        has_row2 = r_label in grid2["row_pos"]
        cells = []
        row_has_data = False
        for c_label in visible_col_headers:
            has1 = has_row1 and c_label in grid1["col_pos"]
            has2 = has_row2 and c_label in grid2["col_pos"]
            v1 = grid1["data"].get((r_label, c_label)) if has1 else None
            v2 = grid2["data"].get((r_label, c_label)) if has2 else None

            if not has1 and not has2:
                status = "empty"
            elif not has1:
                status = "missing1"
                total_diff += 1
            elif not has2:
                status = "missing2"
                total_diff += 1
            elif v1 in (None, "") and v2 not in (None, ""):
                status = "empty_vs_value"
                total_diff += 1
            elif v2 in (None, "") and v1 not in (None, ""):
                status = "empty_vs_value"
                total_diff += 1
            elif v1 == v2:
                status = "same"
            else:
                status = "diff"
                total_diff += 1

            if v1 not in (None, "") or v2 not in (None, ""):
                row_has_data = True

            cells.append({
                "col": c_label,
                "v1": "" if v1 is None else v1,
                "v2": "" if v2 is None else v2,
                "status": status,
            })

        if row_has_data:
            rows_out.append({"row": r_label, "cells": cells})

    return {
        "col_headers": visible_col_headers,
        "display_col_headers": display_col_headers,
        "row_headers": row_headers,
        "rows": rows_out,
        "total_differences": total_diff,
    }


def diff_workbooks_by_header(path1, path2, reference_path=None, reference_grids=None):
    wb1 = load_workbook(path1, data_only=False)
    wb2 = load_workbook(path2, data_only=False)
    reference_wb = load_workbook(reference_path, data_only=False) if reference_path else None

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
    common_sheets = sorted(sheets1 & sheets2)

    sheets_out = {}
    total = 0
    for name in common_sheets:
        reference_grid = None
        if reference_grids is not None and name in reference_grids:
            reference_grid = reference_grids[name]
        elif reference_wb is not None and name in reference_wb.sheetnames:
            reference_grid = build_header_grid(reference_wb[name])

        g1 = build_header_grid(wb1[name], reference_grid=reference_grid)
        g2 = build_header_grid(wb2[name], reference_grid=reference_grid)
        cmp_result = compare_sheet_grids(g1, g2, reference_grid=reference_grid)
        if reference_grid is not None:
            cmp_result["reference_col_headers"] = list(reference_grid.get("col_headers", []))
            cmp_result["reference_col_headers_display"] = list(reference_grid.get("display_col_headers", reference_grid.get("col_headers", [])))
            cmp_result["reference_row_headers"] = list(reference_grid.get("row_headers", []))
        else:
            cmp_result["reference_col_headers"] = list(cmp_result.get("col_headers", []))
            cmp_result["reference_col_headers_display"] = list(cmp_result.get("display_col_headers", cmp_result.get("col_headers", [])))
            cmp_result["reference_row_headers"] = list(cmp_result.get("row_headers", []))
        cmp_result["header_row_used"] = (reference_grid or g1).get("header_row", 1)
        sheets_out[name] = cmp_result
        total += cmp_result["total_differences"]

    return {
        "sheets_only_in_file1": sorted(sheets1 - sheets2),
        "sheets_only_in_file2": sorted(sheets2 - sheets1),
        "sheets": sheets_out,
        "summary": {
            "total_differences": total,
            "sheets_compared": common_sheets,
        },
    }


def build_reference_grids_for_paths(paths, reference_path=None):
    sheet_names = set()
    for path in paths:
        if not os.path.exists(path):
            continue
        try:
            wb = load_workbook(path, data_only=False)
        except Exception:
            continue
        sheet_names.update(wb.sheetnames)

    reference_grids = {}
    for sheet_name in sheet_names:
        reference_grid = None
        if reference_path and os.path.exists(reference_path):
            try:
                reference_wb = load_workbook(reference_path, data_only=False)
            except Exception:
                reference_wb = None
            if reference_wb is not None and sheet_name in reference_wb.sheetnames:
                reference_grid = build_header_grid(reference_wb[sheet_name])
        if reference_grid is None:
            reference_grid = build_reference_grid_from_paths(paths, sheet_name)
        if reference_grid is not None:
            reference_grids[sheet_name] = reference_grid

    return reference_grids


def compare_all(base_path, compare_paths, base_name, compare_names, reference_path=None):
    comparisons = []
    reference_grids = build_reference_grids_for_paths([base_path] + compare_paths, reference_path=reference_path)

    for name, path in zip(compare_names, compare_paths):
        result = diff_workbooks_by_header(base_path, path, reference_path=reference_path, reference_grids=reference_grids)
        result["compare_name"] = name
        comparisons.append(result)

    return {
        "file1_name": base_name,
        "compare_names": compare_names,
        "comparisons": comparisons,
    }


def build_directional_comparisons(file1_path, file2_path, file1_name, file2_name, reference_path=None):
    reference_grids = build_reference_grids_for_paths([file1_path, file2_path], reference_path=reference_path)
    comparisons = []

    for left_path, right_path, left_name, right_name, direction in [
        (file1_path, file2_path, file1_name, file2_name, "file1_to_file2"),
        (file2_path, file1_path, file2_name, file1_name, "file2_to_file1"),
    ]:
        result = diff_workbooks_by_header(left_path, right_path, reference_path=reference_path, reference_grids=reference_grids)
        result["compare_name"] = f"{left_name} → {right_name}"
        result["direction"] = direction
        result["left_name"] = left_name
        result["right_name"] = right_name
        comparisons.append(result)

    return comparisons


@app.after_request
def add_no_store_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.errorhandler(Exception)
def handle_any_exception(e):
    if isinstance(e, HTTPException):
        return jsonify({"error": e.description}), e.code
    return jsonify({"error": str(e)}), 500


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/compare", methods=["POST"])
def compare():
    cleanup_expired_jobs()

    file1 = request.files.get("file1")
    file2 = request.files.get("file2")
    template_file = request.files.get("template")

    if not file1 or not file2:
        return jsonify({"error": "File 1 dan File 2 wajib diunggah."}), 400

    job_id = uuid.uuid4().hex
    job_dir = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)

    path1 = os.path.join(job_dir, "file1.xlsx")
    path2 = os.path.join(job_dir, "file2.xlsx")
    file1.save(path1)
    file2.save(path2)

    try:
        template_path = None
        template_display_name = None

        if template_file and getattr(template_file, "filename", ""):
            # Template diunggah manual untuk sesi ini saja — disimpan di
            # job_dir yang sementara, TIDAK ditambahkan ke library skeleton
            # permanen.
            template_path = os.path.join(job_dir, "template.xlsx")
            template_file.save(template_path)
            template_display_name = template_file.filename

        inferred_template_path = None
        inferred_template_name = None
        for candidate_path, candidate_name in [(path1, file1.filename), (path2, file2.filename)]:
            if inferred_template_path is None and detect_skeleton_from_workbook(candidate_path):
                inferred_template_path = candidate_path
                inferred_template_name = candidate_name

        if template_path is None:
            inferred_from_skeletons = find_matching_skeleton_for_paths([path1, path2])
            if inferred_from_skeletons is not None:
                template_path = inferred_from_skeletons
                template_display_name = os.path.basename(inferred_from_skeletons)
            elif inferred_template_path is not None:
                template_path = inferred_template_path
                template_display_name = inferred_template_name

        comparisons = build_directional_comparisons(path1, path2, file1.filename, file2.filename, reference_path=template_path)

        # Simpan salinan template yang benar-benar dipakai (dari SKELETON_DIR
        # atau dari isi file1/file2 sendiri) ke job_dir yang sementara, supaya
        # /api/merge bisa memakainya. Ini tetap sementara — ikut terhapus
        # otomatis bersama seluruh job_dir.
        if template_path is not None:
            canonical_template_path = os.path.join(job_dir, "template.xlsx")
            if os.path.abspath(template_path) != os.path.abspath(canonical_template_path):
                try:
                    shutil.copyfile(template_path, canonical_template_path)
                except OSError:
                    pass

        touch_job(job_id)

        result = {
            "job_id": job_id,
            "file1_name": file1.filename,
            "file2_name": file2.filename,
            "compare_names": [f"{file1.filename} → {file2.filename}", f"{file2.filename} → {file1.filename}"],
            "comparisons": comparisons,
            "template_detected": template_path is not None,
            "template_name": template_display_name,
            "template_source": "uploaded" if template_file and getattr(template_file, "filename", "") else ("inferred" if template_path else "none"),
            "header_source": "skeleton" if template_path else "shared",
            "header_source_label": "skeleton/template" if template_path else "header umum dari file unggahan",
        }
        return jsonify(result)
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({"error": str(e)}), 500
    # File hasil unggahan disimpan sementara di JOBS_DIR (folder temp OS) agar
    # bisa dipakai fitur "Gabungkan File" (/api/merge). Tidak perlu dihapus
    # manual — cleanup_expired_jobs() menghapusnya otomatis setelah
    # JOB_TTL_SECONDS tanpa aktivitas.


@app.route("/api/merge", methods=["POST"])
def merge():
    cleanup_expired_jobs()

    data = request.get_json(silent=True) or {}
    job_id = data.get("job_id")
    base_file = data.get("base_file", "base")
    choices = data.get("choices", [])

    if not job_id:
        return jsonify({"error": "job_id tidak ada. Silakan bandingkan ulang file."}), 400

    job_dir = os.path.join(JOBS_DIR, job_id)
    # File 1 dan File 2 di sini SELALU dua file asli yang diunggah saat
    # membandingkan — tidak bergantung pada arah/tab perbandingan yang
    # sedang dilihat di UI.
    base_path = os.path.join(job_dir, "file1.xlsx")
    compare_path = os.path.join(job_dir, "file2.xlsx")
    template_path = os.path.join(job_dir, "template.xlsx")

    if not os.path.exists(base_path) or not os.path.exists(compare_path):
        return jsonify({"error": "Sesi sudah kadaluarsa. Silakan unggah dan bandingkan ulang file."}), 400

    if base_file == "template":
        if not os.path.exists(template_path):
            return jsonify({"error": "Template/skeleton tidak tersedia untuk dipakai sebagai kerangka hasil."}), 400
        skeleton_path = template_path
    elif base_file == "file2":
        skeleton_path = compare_path
    else:
        skeleton_path = base_path

    try:
        wb_skeleton = load_workbook(skeleton_path, data_only=False)
        wb_file1_vals = load_workbook(base_path, data_only=False)
        wb_file2_vals = load_workbook(compare_path, data_only=False)

        skeleton_grids = {}
        file1_grids = {}
        file2_grids = {}

        for choice in choices:
            sheet = choice.get("sheet")
            row = choice.get("row")
            col = choice.get("col")
            source = choice.get("source")

            if not sheet or not row or not col or source not in ("file1", "file2"):
                continue
            if sheet not in wb_skeleton.sheetnames:
                continue

            if sheet not in skeleton_grids:
                skeleton_grids[sheet] = build_header_grid(wb_skeleton[sheet])
            skel_grid = skeleton_grids[sheet]

            if row not in skel_grid["row_pos"] or col not in skel_grid["col_pos"]:
                continue

            target_row = skel_grid["row_pos"][row]
            target_col = skel_grid["col_pos"][col]

            if source == "file1":
                if sheet not in wb_file1_vals.sheetnames:
                    continue
                if sheet not in file1_grids:
                    file1_grids[sheet] = build_header_grid(wb_file1_vals[sheet], reference_grid=skel_grid)
                src_grid = file1_grids[sheet]
            else:
                if sheet not in wb_file2_vals.sheetnames:
                    continue
                if sheet not in file2_grids:
                    file2_grids[sheet] = build_header_grid(wb_file2_vals[sheet], reference_grid=skel_grid)
                src_grid = file2_grids[sheet]

            if row not in src_grid["row_pos"] or col not in src_grid["col_pos"]:
                new_value = None
            else:
                new_value = src_grid["data"].get((row, col))

            wb_skeleton[sheet].cell(row=target_row, column=target_col).value = new_value

        buffer = io.BytesIO()
        wb_skeleton.save(buffer)
        buffer.seek(0)

        # Sesi masih aktif dipakai — perpanjang umurnya supaya tidak terhapus
        # otomatis kalau pengguna ingin mencoba pilihan merge lain.
        touch_job(job_id)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="file_gabungan.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True, port=5000)